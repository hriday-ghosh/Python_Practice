# OpenPyXL Basic and Advanced Syntaxes Explained Simply

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------- BASIC ----------------

# 1. Create a New Workbook
wb = Workbook()  # Create a blank workbook
ws = wb.active  # Get the default active sheet
ws.title = "MySheet"  # Rename the sheet

# 2. Writing Data to Cells
ws['A1'] = "Name"  # Set cell A1
ws['B1'] = "Score"  # Set cell B1
ws['A2'] = "Alice"
ws['B2'] = 90

# 3. Save the Workbook
wb.save("sample.xlsx")  # Save to file

# 4. Load Existing Workbook
wb2 = load_workbook("sample.xlsx")  # Load an existing file
ws2 = wb2.active  # Get the active sheet again

# 5. Read Data from Cells
print(ws2['A2'].value)  # Print value from A2
print(ws2.cell(row=2, column=2).value)  # Same using row/column method

# 6. Append Rows
ws2.append(["Bob", 85])  # Appends to next empty row
wb2.save("sample.xlsx")  # Save the updated file

# 7. Loop Through Data
for row in ws2.iter_rows(min_row=1, max_row=3, max_col=2):
    for cell in row:
        print(cell.value)  # Print each cell value

# ---------------- ADVANCED ----------------

# 8. Apply Font Style
ws2['A1'].font = Font(bold=True, color="FF0000")  # Bold red font

# 9. Alignment
ws2['A1'].alignment = Alignment(horizontal="center")  # Center text

# 10. Fill Color
ws2['A1'].fill = PatternFill(start_color="FFFF00", fill_type="solid")  # Yellow background

# 11. Merge Cells
ws2.merge_cells("C1:D1")  # Merge C1 and D1
ws2["C1"] = "Merged Cell"

# 12. Freeze Panes
ws2.freeze_panes = "B2"  # Freeze first row and column

# 13. Auto Filter
ws2.auto_filter.ref = "A1:B10"  # Apply filter to range

# 14. Column Width
ws2.column_dimensions['A'].width = 20  # Make column A wider

# 15. Delete Rows/Columns
ws2.delete_rows(4)  # Delete 4th row
ws2.delete_cols(3)  # Delete 3rd column

# 16. Insert Rows/Columns
ws2.insert_rows(4)  # Insert a row at position 4
ws2.insert_cols(3)  # Insert a column at position 3

# 17. Save Changes
wb2.save("sample.xlsx")  # Save again after changes

# You can now read, write, format, and manipulate Excel using OpenPyXL!
